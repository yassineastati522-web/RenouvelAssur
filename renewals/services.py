import csv
import re
import unicodedata
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import StringIO
from zipfile import BadZipFile

from django.conf import settings
from django.db import transaction
from django.db.models.functions import Lower
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .models import Client, Contract, ImportBatch, Termination


# Les alias sont ordonnés : lorsqu'un fichier contient deux colonnes équivalentes,
# la première valeur non vide est utilisée. Par exemple, IMMATDEF est prioritaire
# sur IMMAPRO dans les bordereaux fournis par l'assureur.
ALIASES = {
    "category": ("cat", "categorie", "type contrat"),
    "policy_number": ("police", "numero police", "n police"),
    "agent_reference": ("reference agent", "ref agent", "libelle intermediaire"),
    "agent_code": ("code agent", "code intermediaire"),
    "event": ("nature evenement", "nature evennement", "evenement", "event"),
    "renewed_flag": ("renouvele", "renouvellement", "statut renouvellement"),
    "pack_code": ("code pack convention", "code pack", "convention"),
    "client_name": ("client", "assure", "nom assure", "nom client"),
    "client_phone": ("telephone", "telephone2", "telephone 2", "tel", "mobile", "numero telephone"),
    "client_external_id": ("numero cin", "cin", "identifiant client", "id client", "code client"),
    "brand": ("marque", "marque vehicule"),
    "registration": ("immatdef", "immatriculation definitive", "immatriculation", "matricule", "immapro", "immatriculation provisoire"),
    "net_premium": ("prime net", "prime nette"),
    "cash_premium": ("prime au comptant",),
    "total_premium": ("prime total", "prime totale", "prime ttc"),
    "net_payable": ("net a paye", "net a payer"),
    "receipt": ("n quittance", "num quittance", "numero quittance", "quittance"),
    "effective_date": ("date effet", "date d effet", "date debut"),
    "end_date": ("date echeance", "date fin", "date de fin", "date fin echeance"),
    "issue_date": ("date emission", "date d emission"),
    "provisional_attestation": (
        "n attestation",
        "numero attestation",
        "attestation provisoire",
    ),
    "provisional_due_date": (
        "date d echeance",
        "echeance provisoire",
    ),
    "provisional_delivered_count": (
        "provisoires delivrees",
        "nombre provisoires delivrees",
    ),
    "provisional_status": ("etat contrat", "etat du contrat"),
}

REQUIRED_CONTRACT_FIELDS = {"policy_number", "client_name", "end_date"}
SUMMARY_PREFIXES = ("nombre total", "total ht", "total ttc", "total general", "sous total")
UPCOMING_SIGNATURE_FIELDS = {"category", "policy_number", "client_name", "effective_date", "end_date", "renewed_flag"}
BORDEREAU_SIGNATURE_FIELDS = {"event", "receipt", "total_premium"}
PROVISIONAL_SIGNATURE_FIELDS = {
    "policy_number",
    "client_name",
    "registration",
    "effective_date",
    "end_date",
    "provisional_attestation",
    "provisional_due_date",
    "provisional_delivered_count",
}


def normalize(value):
    value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def clean_text(value):
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_policy(value):
    return re.sub(r"\s*/\s*", "/", clean_text(value))


def canonical_policy(value, category=""):
    policy = clean_policy(value)
    category = clean_text(category)
    if category and policy and "/" not in policy:
        return f"{category}/{policy}"
    return policy


def normalized_identity(value):
    return normalize(value).replace(" ", "")


def is_termination_event(value):
    """Indique si une ligne représente un événement qui arrête le contrat."""
    event = normalize(value)
    return bool(event) and any(
        normalize(token) in event
        for token in settings.TERMINATION_EVENTS
        if normalize(token)
    )


def detect_import_type(mapping):
    fields = set(mapping)
    if PROVISIONAL_SIGNATURE_FIELDS <= fields:
        return ImportBatch.ImportType.PROVISIONAL
    if UPCOMING_SIGNATURE_FIELDS <= fields:
        return ImportBatch.ImportType.UPCOMING
    if fields & BORDEREAU_SIGNATURE_FIELDS:
        return ImportBatch.ImportType.BORDEREAU
    if (
        "end_date" not in fields
        and "client_phone" in fields
        and fields & {"policy_number", "client_external_id", "client_name"}
    ):
        return ImportBatch.ImportType.CONTACTS
    return ImportBatch.ImportType.GENERAL


def header_map(headers):
    normalized = {}
    for index, header in enumerate(headers or ()):
        key = normalize(header)
        if key:
            normalized.setdefault(key, []).append(index)

    result = {}
    for field, aliases in ALIASES.items():
        indices = []
        for alias in aliases:
            indices.extend(normalized.get(alias, []))
        if indices:
            result[field] = indices
    return result


def row_value(row, mapping, field):
    indices = mapping.get(field, ())
    for index in indices:
        if index < len(row) and row[index] not in (None, ""):
            return row[index]
    return ""


def find_header(rows, scan_limit=30):
    best_index, best_mapping, best_rank = 0, {}, (-1, -1, 0)
    for index, row in enumerate(rows[:scan_limit]):
        mapping = header_map(row)
        rank = (len(REQUIRED_CONTRACT_FIELDS & mapping.keys()), len(mapping), -index)
        if rank > best_rank:
            best_index, best_mapping, best_rank = index, mapping, rank
    return best_index, best_mapping, best_rank


def read_csv_rows(upload):
    upload.seek(0)
    content = upload.read()
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
        except UnicodeDecodeError:
            continue
        if text.strip():
            break
    else:
        raise ValueError("Impossible de décoder ce fichier CSV.")

    first_line = next((line for line in text.splitlines() if line.strip()), "")
    delimiter = ";" if first_line.count(";") >= first_line.count(",") else ","
    rows = list(csv.reader(StringIO(text), delimiter=delimiter))
    header_index, _mapping, _rank = find_header(rows)
    return rows[header_index:] if rows else []


def read_rows(upload):
    """Lit le tableau CSV/Excel pertinent et le recadre sur ses en-têtes."""
    filename = upload.name.lower()
    if filename.endswith(".csv"):
        return read_csv_rows(upload)
    if not filename.endswith((".xlsx", ".xls")):
        raise ValueError(
            "Seuls les fichiers .csv de suivi provisoire et les fichiers "
            "Excel .xlsx ou .xls sont acceptés."
        )

    upload.seek(0)
    try:
        workbook = load_workbook(upload, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise ValueError(
            "Impossible de lire ce fichier Excel. Utilisez un fichier .xlsx ou le fichier .xls "
            "fourni par l’assureur."
        ) from exc

    try:
        best_rows, best_header_index, best_rank = [], 0, (-1, -1, 0, 0)
        for sheet_index, sheet in enumerate(workbook.worksheets):
            rows = list(sheet.iter_rows(values_only=True))
            header_index, _mapping, rank = find_header(rows)
            sheet_rank = (*rank[:2], -sheet_index, rank[2])
            if sheet_rank > best_rank:
                best_rows, best_header_index, best_rank = rows, header_index, sheet_rank
        return best_rows[best_header_index:] if best_rows else []
    finally:
        workbook.close()


def parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"date invalide : {text}")


def parse_contract_end_date(value, import_type):
    """Convertit une borne d'échéance à minuit en dernier jour couvert.

    Les bordereaux de production et les suivis provisoires expriment la fin
    sous forme de borne exclusive (par exemple 01/08/2026 à 00:00 pour une
    couverture qui se termine le 31/07/2026). Le fichier des échéances à
    venir contient déjà le dernier jour couvert et ne doit pas être décalé.
    """
    parsed = parse_date(value)
    if parsed and import_type in {
        ImportBatch.ImportType.BORDEREAU,
        ImportBatch.ImportType.PROVISIONAL,
    }:
        return parsed - timedelta(days=1)
    return parsed


def parse_decimal(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation as exc:
        raise ValueError(f"montant invalide : {value}") from exc


def parse_provisional_count(value):
    text = clean_text(value)
    try:
        count = int(text)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"nombre de provisoires invalide : {text or 'vide'}") from exc
    if count < 1 or count > 3:
        raise ValueError("le nombre de provisoires délivrées doit être compris entre 1 et 3")
    return count


def provisional_quota(effective_date, end_date):
    if not effective_date or not end_date or end_date <= effective_date:
        raise ValueError("dates du contrat invalides pour calculer les provisoires")
    duration_days = (end_date - effective_date).days
    if duration_days <= 120:
        return 1
    if duration_days <= 240:
        return 2
    return 3


def clean_provisional_status(value):
    return clean_text(value).split("?", 1)[0].strip()


def provisional_is_active(status):
    normalized = normalize(status)
    return normalized not in {
        "sans effet",
        "annule",
        "annulee",
        "termine",
        "terminee",
    }


def is_summary_row(row, mapping):
    if row_value(row, mapping, "client_name") or row_value(row, mapping, "end_date"):
        return False
    first_value = next((normalize(value) for value in row if value not in (None, "")), "")
    return any(first_value.startswith(prefix) for prefix in SUMMARY_PREFIXES)


def data_rows(rows, mapping):
    for line_number, row in enumerate(rows[1:], 2):
        if not any(value not in (None, "") for value in row):
            continue
        if is_summary_row(row, mapping):
            continue
        yield line_number, row


def analyze_rows(rows):
    analysis = {
        "total_rows": 0,
        "recognized": [],
        "errors": [],
        "valid": False,
        "import_type": ImportBatch.ImportType.GENERAL,
    }
    if not rows:
        analysis["errors"].append({"line": 1, "error": "Fichier vide"})
        return analysis

    mapping = header_map(rows[0])
    analysis["import_type"] = detect_import_type(mapping)
    analysis["recognized"] = [
        clean_text(rows[0][index])
        for index in sorted({index for indices in mapping.values() for index in indices})
    ]
    candidates = list(data_rows(rows, mapping))
    analysis["total_rows"] = len(candidates)
    contacts_only = analysis["import_type"] == ImportBatch.ImportType.CONTACTS
    missing = REQUIRED_CONTRACT_FIELDS - mapping.keys()
    if missing and not contacts_only:
        analysis["errors"].append({"line": 1, "error": "Colonnes obligatoires absentes : " + ", ".join(sorted(missing))})
        return analysis

    for line_number, row in candidates:
        try:
            if contacts_only:
                if not clean_text(row_value(row, mapping, "client_phone")):
                    raise ValueError("téléphone obligatoire")
            else:
                if not clean_text(row_value(row, mapping, "policy_number")) or not clean_text(row_value(row, mapping, "client_name")):
                    raise ValueError("police et assuré obligatoires")
                end_date = parse_contract_end_date(
                    row_value(row, mapping, "end_date"),
                    analysis["import_type"],
                )
                if not end_date:
                    raise ValueError("date de fin obligatoire")
                effective_date = parse_date(
                    row_value(row, mapping, "effective_date")
                )
                for field in ("effective_date", "issue_date"):
                    parse_date(row_value(row, mapping, field))
                for field in ("net_premium", "cash_premium", "total_premium", "net_payable"):
                    parse_decimal(row_value(row, mapping, field))
                if analysis["import_type"] == ImportBatch.ImportType.PROVISIONAL:
                    due_date = parse_date(
                        row_value(row, mapping, "provisional_due_date")
                    )
                    delivered_count = parse_provisional_count(
                        row_value(row, mapping, "provisional_delivered_count")
                    )
                    allowed_count = provisional_quota(effective_date, end_date)
                    if not due_date:
                        raise ValueError("date d’échéance provisoire obligatoire")
                    if not effective_date <= due_date <= end_date:
                        raise ValueError(
                            "l’échéance provisoire doit être comprise dans "
                            "la période du contrat"
                        )
                    if delivered_count > allowed_count:
                        raise ValueError(
                            f"{delivered_count} provisoires délivrées alors que "
                            f"ce contrat en autorise {allowed_count}"
                        )
        except Exception as exc:
            if len(analysis["errors"]) < 100:
                analysis["errors"].append({"line": line_number, "error": str(exc)})
    analysis["valid"] = not any(error["line"] == 1 for error in analysis["errors"])
    return analysis


def record_error(batch, line_number, exc):
    batch.rejected_rows += 1
    if len(batch.errors) < 100:
        batch.errors.append({"line": line_number, "error": str(exc)})


def save_batch(batch):
    batch.save(update_fields=["added_rows", "updated_rows", "rejected_rows", "errors"])
    return batch


def import_contact_rows(candidates, mapping, batch):
    parsed = []
    for line_number, row in candidates:
        try:
            phone = clean_text(row_value(row, mapping, "client_phone"))
            if not phone:
                raise ValueError("téléphone obligatoire")
            parsed.append({
                "line": line_number,
                "phone": phone,
                "external_id": clean_text(row_value(row, mapping, "client_external_id")),
                "policy": clean_text(row_value(row, mapping, "policy_number")),
                "name": clean_text(row_value(row, mapping, "client_name")),
            })
        except Exception as exc:
            record_error(batch, line_number, exc)

    external_ids = {item["external_id"] for item in parsed if item["external_id"]}
    policies = {item["policy"] for item in parsed if item["policy"]}
    names = {item["name"].lower() for item in parsed if item["name"]}
    clients_by_external_id = {
        client.external_id: client
        for client in Client.objects.filter(external_id__in=external_ids)
    }
    clients_by_policy = {
        contract.policy_number: contract.client
        for contract in Contract.objects.select_related("client").filter(policy_number__in=policies)
    }
    clients_by_name = {}
    if names:
        for client in Client.objects.annotate(import_name=Lower("name")).filter(import_name__in=names):
            clients_by_name.setdefault(client.name.lower(), client)

    changed_clients = {}
    for item in parsed:
        client = None
        if item["external_id"]:
            client = clients_by_external_id.get(item["external_id"])
        if client is None and item["policy"]:
            client = clients_by_policy.get(item["policy"])
        if client is None and item["name"]:
            client = clients_by_name.get(item["name"].lower())
        if client is None:
            record_error(batch, item["line"], ValueError("client introuvable avec les identifiants fournis"))
            continue
        client.phone = item["phone"]
        changed_clients[client.pk] = client
        batch.updated_rows += 1

    if changed_clients:
        now = timezone.now()
        for client in changed_clients.values():
            client.updated_at = now
        Client.objects.bulk_update(
            changed_clients.values(),
            ["phone", "updated_at"],
            batch_size=500,
        )
    return save_batch(batch)


CONTRACT_VALUE_FIELDS = (
    "category", "agent_reference", "agent_code", "event", "pack_code", "brand",
    "registration", "net_premium", "cash_premium", "total_premium", "net_payable",
    "effective_date", "end_date", "issue_date", "is_provisional",
    "provisional_attestation", "provisional_due_date",
    "provisional_delivered_count", "provisional_allowed_count",
    "provisional_selected_count", "provisional_status",
)


def policy_from_row(row, mapping, import_type):
    policy = clean_policy(row_value(row, mapping, "policy_number"))
    category = clean_text(row_value(row, mapping, "category"))
    if import_type == ImportBatch.ImportType.UPCOMING:
        return canonical_policy(policy, category)
    return policy


def item_fingerprint(item):
    values = item["values"]
    return (
        canonical_policy(item["policy"], values["category"]),
        normalized_identity(item["name"]),
        normalized_identity(values["registration"]),
        values["effective_date"],
        values["end_date"],
        normalize(values["event"]),
        values["total_premium"],
        values["net_premium"],
        values.get("provisional_attestation", ""),
        values.get("provisional_due_date"),
        values.get("provisional_delivered_count", 0),
    )


def contract_fingerprint(contract):
    return (
        canonical_policy(contract.policy_number, contract.category),
        normalized_identity(contract.client.name),
        normalized_identity(contract.registration),
        contract.effective_date,
        contract.end_date,
        normalize(contract.event),
        contract.total_premium,
        contract.net_premium,
        contract.provisional_attestation,
        contract.provisional_due_date,
        contract.provisional_delivered_count,
    )


def contract_business_identity(
    policy,
    client_name,
    registration,
    effective_date,
    end_date,
    category="",
):
    vehicle = normalized_identity(registration)
    if not vehicle:
        return None
    return (
        canonical_policy(policy, category),
        normalized_identity(client_name),
        vehicle,
        effective_date,
        end_date,
    )


def item_business_identity(item):
    values = item["values"]
    return contract_business_identity(
        item["policy"],
        item["name"],
        values["registration"],
        values["effective_date"],
        values["end_date"],
        values["category"],
    )


def stored_contract_business_identity(contract):
    return contract_business_identity(
        contract.policy_number,
        contract.client.name,
        contract.registration,
        contract.effective_date,
        contract.end_date,
        contract.category,
    )


def premium_rank(value):
    return value if value is not None else Decimal("-Infinity")


def negative_premium(value):
    if value in (None, 0):
        return None
    return -abs(value)


def indicates_renewed(value):
    flag = normalize(value)
    if not flag or flag.startswith("non ") or flag in {"non", "no", "0", "false"}:
        return False
    return flag in {"oui", "yes", "1", "true"} or "renouvel" in flag


def dates_match(left, right, tolerance_days=1):
    return bool(left and right and abs((left - right).days) <= tolerance_days)


def same_contract_subject(item, contract):
    """Rapproche une résiliation sans dépendre de sa quittance/date d'effet."""
    if canonical_policy(
        item["policy"],
        item["values"]["category"],
    ) != canonical_policy(contract.policy_number, contract.category):
        return False

    incoming_external_id = normalized_identity(item["external_id"])
    stored_external_id = normalized_identity(contract.client.external_id)
    if incoming_external_id and stored_external_id:
        if incoming_external_id != stored_external_id:
            return False
    elif normalized_identity(item["name"]) != normalized_identity(
        contract.client.name
    ):
        return False

    incoming_vehicle = normalized_identity(item["values"]["registration"])
    stored_vehicle = normalized_identity(contract.registration)
    if incoming_vehicle and stored_vehicle and incoming_vehicle != stored_vehicle:
        return False
    return True


def select_termination_candidate(item, contracts):
    """Trouve le contrat actif visé par une ligne de résiliation."""
    cancellation_date = item["termination_date"]
    source_end_date = item["source_end_date"]
    candidates = []
    for contract in contracts:
        if not same_contract_subject(item, contract):
            continue
        already_closed = (
            contract.manually_terminated
            or contract.renewal_status == Contract.RenewalStatus.TERMINATED
            or hasattr(contract, "termination")
        )
        exact_expiry = dates_match(contract.end_date, source_end_date)
        same_recorded_stop = bool(
            hasattr(contract, "termination")
            and contract.termination.date == cancellation_date
        )
        same_closed_cycle = bool(
            already_closed
            and source_end_date
            and contract.end_date <= source_end_date + timedelta(days=1)
            and (
                not contract.effective_date
                or contract.effective_date <= contract.end_date
            )
        )
        cancellation_in_period = bool(
            cancellation_date
            and cancellation_date <= contract.end_date + timedelta(days=1)
            and (
                not contract.effective_date
                or cancellation_date >= contract.effective_date
            )
        )
        if not exact_expiry and not cancellation_in_period and not same_closed_cycle:
            continue
        if (
            not same_recorded_stop
            and not exact_expiry
            and not same_closed_cycle
            and contract.effective_date
            and contract.effective_date >= cancellation_date
        ):
            # Un contrat qui commence au moment de l'arrêt peut être son
            # successeur légitime ; il ne faut pas le résilier par erreur.
            continue
        expiry_distance = (
            abs((contract.end_date - source_end_date).days)
            if source_end_date
            else 10**9
        )
        candidates.append(
            (
                not same_recorded_stop,
                not exact_expiry,
                not same_closed_cycle,
                already_closed,
                expiry_distance,
                contract.policy_number != item["policy"],
                contract.pk,
                contract,
            )
        )
    return min(candidates, default=(None,) * 8)[-1]


def indexed_policy_candidates(index, *aliases):
    candidates = {}
    for alias in aliases:
        for contract in index.get(alias, ()):
            candidates[contract.pk] = contract
    return list(candidates.values())


def select_closed_cycle_candidate(item, contracts):
    """Évite qu'un réimport d'échéances recrée un contrat déjà résilié."""
    incoming_start = item["values"]["effective_date"]
    incoming_end = item["source_end_date"]
    if not incoming_end:
        return None
    candidates = []
    for contract in contracts:
        already_closed = (
            contract.manually_terminated
            or contract.renewal_status == Contract.RenewalStatus.TERMINATED
            or hasattr(contract, "termination")
        )
        if not already_closed or not same_contract_subject(item, contract):
            continue
        stopped_on = contract.end_date
        if incoming_start and stopped_on < incoming_start:
            continue
        if stopped_on > incoming_end + timedelta(days=1):
            continue
        candidates.append((stopped_on, contract.pk, contract))
    return min(candidates, default=(None,) * 3)[-1]


def same_import_subject(left, right):
    if canonical_policy(
        left["policy"],
        left["values"]["category"],
    ) != canonical_policy(
        right["policy"],
        right["values"]["category"],
    ):
        return False
    if normalized_identity(left["name"]) != normalized_identity(right["name"]):
        return False
    left_vehicle = normalized_identity(left["values"]["registration"])
    right_vehicle = normalized_identity(right["values"]["registration"])
    if left_vehicle and right_vehicle and left_vehicle != right_vehicle:
        return False
    return dates_match(left["source_end_date"], right["source_end_date"])


def combine_batch_termination_rows(items):
    """Fusionne production + résiliation du même contrat dans un seul objet."""
    regular = [item for item in items if not item["is_termination"]]
    terminations = [item for item in items if item["is_termination"]]
    consumed_lines = set()
    combined = []

    termination_groups = []
    grouped_lines = set()
    for termination_item in terminations:
        if termination_item["line"] in grouped_lines:
            continue
        group = [
            item
            for item in terminations
            if item["line"] not in grouped_lines
            and same_import_subject(termination_item, item)
        ]
        grouped_lines.update(item["line"] for item in group)
        termination_groups.append(group)

    for termination_group in termination_groups:
        termination_item = min(
            termination_group,
            key=lambda item: (item["termination_date"], item["line"]),
        )
        candidates = [
            item
            for item in regular
            if item["line"] not in consumed_lines
            and same_import_subject(termination_item, item)
        ]
        if not candidates:
            combined.append(termination_item)
            continue
        base = max(
            candidates,
            key=lambda item: (
                premium_rank(item["values"]["total_premium"]),
                bool(item["receipt"]),
                -item["line"],
            ),
        )
        consumed_lines.add(base["line"])
        merged = dict(base)
        merged_values = dict(base["values"])
        for field, value in termination_item["values"].items():
            if field == "event" or merged_values.get(field) in (None, ""):
                merged_values[field] = value
        merged.update({
            "line": termination_item["line"],
            "is_termination": True,
            "termination_date": termination_item["termination_date"],
            "termination_reason": termination_item["termination_reason"],
            "termination_premium": termination_item["termination_premium"],
            "source_end_date": termination_item["source_end_date"],
            "renewed": False,
            "values": merged_values,
        })
        combined.append(merged)

    combined.extend(
        item for item in regular if item["line"] not in consumed_lines
    )
    return sorted(combined, key=lambda item: item["line"])


def select_contract_candidate(item, contracts, claimed_ids):
    candidates = [
        contract
        for contract in contracts
        if contract.pk not in claimed_ids
        and dates_match(contract.end_date, item["values"]["end_date"])
    ]
    if item["import_type"] == ImportBatch.ImportType.BORDEREAU:
        candidates = [
            contract
            for contract in candidates
            if not contract.receipt or contract.from_upcoming_file
        ]
    if not candidates:
        return None
    incoming_end = item["values"]["end_date"]
    return min(
        candidates,
        key=lambda contract: (
            abs((contract.end_date - incoming_end).days),
            contract.policy_number != item["policy"],
            not contract.from_upcoming_file,
            bool(contract.receipt),
            contract.pk,
        ),
    )


def merge_contract_values(contract, item, client):
    incoming = item["values"]
    preserve_snapshot_end = (
        item["import_type"] == ImportBatch.ImportType.BORDEREAU
        and contract.from_upcoming_file
        and dates_match(contract.end_date, incoming["end_date"])
    )
    contract.client = client
    for field, value in incoming.items():
        if value in (None, ""):
            continue
        if field == "end_date" and preserve_snapshot_end:
            continue
        setattr(contract, field, value)
    if (
        item["import_type"] == ImportBatch.ImportType.PROVISIONAL
        and contract.provisional_selected_count
        and incoming["provisional_delivered_count"]
        > contract.provisional_selected_count
    ):
        contract.provisional_selected_count = incoming[
            "provisional_delivered_count"
        ]
    if item["receipt"]:
        contract.receipt = item["receipt"]
    if item["import_type"] == ImportBatch.ImportType.UPCOMING:
        contract.policy_number = item["policy"]
        contract.from_upcoming_file = True


def merge_termination_values(contract, item, client):
    """Clôture le contrat existant sans remplacer ses données de production."""
    incoming = item["values"]
    contract.client = client
    fill_if_empty = (
        "category",
        "agent_reference",
        "agent_code",
        "pack_code",
        "brand",
        "registration",
        "net_premium",
        "cash_premium",
        "total_premium",
        "net_payable",
        "effective_date",
        "issue_date",
    )
    for field in fill_if_empty:
        value = incoming.get(field)
        if getattr(contract, field) in (None, "") and value not in (None, ""):
            setattr(contract, field, value)
    contract.event = item["termination_reason"]
    contract.end_date = item["termination_date"]
    if not contract.receipt and item["receipt"]:
        contract.receipt = item["receipt"]


def merge_closed_cycle_values(contract, item, client):
    """Enrichit un contrat clôturé sans le rouvrir ni changer son arrêt."""
    incoming = item["values"]
    contract.client = client
    for field in (
        "category",
        "agent_reference",
        "agent_code",
        "pack_code",
        "brand",
        "registration",
    ):
        value = incoming.get(field)
        if getattr(contract, field) in (None, "") and value not in (None, ""):
            setattr(contract, field, value)


def mark_vehicle_renewals():
    contracts = list(
        Contract.objects.exclude(registration="").only(
            "pk",
            "registration",
            "effective_date",
            "issue_date",
            "end_date",
            "renewal_status",
            "renewed_contract_id",
            "manually_terminated",
            "updated_at",
        )
    )
    contracts_by_vehicle = {}
    for contract in contracts:
        vehicle_key = normalized_identity(contract.registration)
        if vehicle_key:
            contracts_by_vehicle.setdefault(vehicle_key, []).append(contract)

    changed = {}
    now = timezone.now()
    claimed_successors = {
        contract.renewed_contract_id: contract.pk
        for contract in contracts
        if contract.renewed_contract_id
    }
    for vehicle_contracts in contracts_by_vehicle.values():
        for old_contract in sorted(
            vehicle_contracts,
            key=lambda contract: (contract.end_date, contract.pk),
        ):
            if (
                old_contract.manually_terminated
                or old_contract.renewal_status
                == Contract.RenewalStatus.TERMINATED
            ):
                continue
            earliest_start = old_contract.end_date - timedelta(days=1)
            successors = []
            for candidate in vehicle_contracts:
                candidate_start = (
                    candidate.effective_date or candidate.issue_date
                )
                if (
                    candidate.pk != old_contract.pk
                    and candidate_start
                    and candidate_start >= earliest_start
                    and candidate.end_date > old_contract.end_date
                    and not candidate.manually_terminated
                    and candidate.renewal_status
                    != Contract.RenewalStatus.TERMINATED
                    and claimed_successors.get(candidate.pk)
                    in (None, old_contract.pk)
                ):
                    successors.append((candidate_start, candidate.end_date, candidate.pk, candidate))
            if not successors:
                continue
            successor = min(successors, key=lambda item: item[:3])[3]
            if (
                old_contract.renewed_contract_id != successor.pk
                or old_contract.renewal_status
                != Contract.RenewalStatus.RENEWED
            ):
                if (
                    old_contract.renewed_contract_id
                    and claimed_successors.get(old_contract.renewed_contract_id)
                    == old_contract.pk
                ):
                    claimed_successors.pop(old_contract.renewed_contract_id)
                old_contract.renewed_contract = successor
                old_contract.renewal_status = Contract.RenewalStatus.RENEWED
                old_contract.updated_at = now
                claimed_successors[successor.pk] = old_contract.pk
                changed[old_contract.pk] = old_contract

    if changed:
        Contract.objects.bulk_update(
            changed.values(),
            ["renewed_contract", "renewal_status", "updated_at"],
            batch_size=500,
        )
    return len(changed)


def import_contract_rows(rows, filename, user):
    mapping = header_map(rows[0]) if rows else {}
    candidates = list(data_rows(rows, mapping)) if rows else []
    import_type = detect_import_type(mapping)
    batch = ImportBatch.objects.create(
        filename=filename,
        import_type=import_type,
        imported_by=user,
        total_rows=len(candidates),
    )
    if not rows:
        batch.errors = [{"line": 1, "error": "Fichier vide"}]
        batch.rejected_rows = 1
        return save_batch(batch)

    contacts_only = import_type == ImportBatch.ImportType.CONTACTS
    missing = REQUIRED_CONTRACT_FIELDS - mapping.keys()
    if missing and not contacts_only:
        batch.errors = [{"line": 1, "error": "Colonnes obligatoires absentes : " + ", ".join(sorted(missing))}]
        batch.rejected_rows = batch.total_rows
        return save_batch(batch)
    if contacts_only:
        return import_contact_rows(candidates, mapping, batch)

    parsed = []
    for line_number, row in candidates:
        try:
            policy = policy_from_row(row, mapping, import_type)
            receipt = clean_text(row_value(row, mapping, "receipt"))
            name = clean_text(row_value(row, mapping, "client_name"))
            if not policy or not name:
                raise ValueError("police et assuré obligatoires")
            values = {
                "category": clean_text(row_value(row, mapping, "category")),
                "agent_reference": clean_text(row_value(row, mapping, "agent_reference")),
                "agent_code": clean_text(row_value(row, mapping, "agent_code")),
                "event": clean_text(row_value(row, mapping, "event")),
                "pack_code": clean_text(row_value(row, mapping, "pack_code")),
                "brand": clean_text(row_value(row, mapping, "brand")),
                "registration": clean_text(row_value(row, mapping, "registration")),
                "net_premium": parse_decimal(row_value(row, mapping, "net_premium")),
                "cash_premium": parse_decimal(row_value(row, mapping, "cash_premium")),
                "total_premium": parse_decimal(row_value(row, mapping, "total_premium")),
                "net_payable": parse_decimal(row_value(row, mapping, "net_payable")),
                "effective_date": parse_date(row_value(row, mapping, "effective_date")),
                "end_date": parse_contract_end_date(
                    row_value(row, mapping, "end_date"),
                    import_type,
                ),
                "issue_date": parse_date(row_value(row, mapping, "issue_date")),
            }
            if not values["end_date"]:
                raise ValueError("date de fin obligatoire")
            if import_type == ImportBatch.ImportType.PROVISIONAL:
                due_date = parse_date(
                    row_value(row, mapping, "provisional_due_date")
                )
                delivered_count = parse_provisional_count(
                    row_value(row, mapping, "provisional_delivered_count")
                )
                allowed_count = provisional_quota(
                    values["effective_date"],
                    values["end_date"],
                )
                if not due_date:
                    raise ValueError("date d’échéance provisoire obligatoire")
                if not values["effective_date"] <= due_date <= values["end_date"]:
                    raise ValueError(
                        "l’échéance provisoire doit être comprise dans "
                        "la période du contrat"
                    )
                if delivered_count > allowed_count:
                    raise ValueError(
                        f"{delivered_count} provisoires délivrées alors que "
                        f"ce contrat en autorise {allowed_count}"
                    )
                provisional_status = clean_provisional_status(
                    row_value(row, mapping, "provisional_status")
                )
                values.update({
                    "is_provisional": provisional_is_active(
                        provisional_status
                    ),
                    "provisional_attestation": clean_text(
                        row_value(row, mapping, "provisional_attestation")
                    ),
                    "provisional_due_date": due_date,
                    "provisional_delivered_count": delivered_count,
                    "provisional_allowed_count": allowed_count,
                    "provisional_status": provisional_status,
                })
            termination_event = is_termination_event(values["event"])
            termination_date = (
                values["effective_date"]
                or values["issue_date"]
                or values["end_date"]
            )
            parsed.append({
                "line": line_number,
                "key": (policy, receipt),
                "policy": policy,
                "legacy_policy": clean_policy(row_value(row, mapping, "policy_number")),
                "receipt": receipt,
                "name": name,
                "external_id": clean_text(row_value(row, mapping, "client_external_id")),
                "phone": clean_text(row_value(row, mapping, "client_phone")),
                "import_type": import_type,
                "renewed": indicates_renewed(row_value(row, mapping, "renewed_flag")),
                "is_termination": termination_event,
                "termination_date": termination_date,
                "termination_reason": values["event"],
                "termination_premium": (
                    negative_premium(values["total_premium"])
                    if termination_event
                    else None
                ),
                "source_end_date": values["end_date"],
                "values": values,
            })
        except Exception as exc:
            record_error(batch, line_number, exc)

    if not parsed:
        return save_batch(batch)

    retained_by_premium = []
    business_groups = {}
    for item in parsed:
        identity = item_business_identity(item)
        if identity is None:
            retained_by_premium.append(item)
        else:
            business_groups.setdefault(identity, []).append(item)

    for group in business_groups.values():
        termination_items = [item for item in group if item["is_termination"]]
        regular_items = [item for item in group if not item["is_termination"]]
        # Une résiliation est un événement de clôture, pas un doublon financier.
        # Sa prime est souvent négative et ne doit jamais la faire rejeter.
        retained_by_premium.extend(termination_items)
        if not regular_items:
            continue
        highest_premium = max(
            premium_rank(item["values"]["total_premium"])
            for item in regular_items
        )
        highest_items = [
            item
            for item in regular_items
            if premium_rank(item["values"]["total_premium"])
            == highest_premium
        ]
        winner_line = min(item["line"] for item in highest_items)
        retained_by_premium.extend(highest_items)
        for item in regular_items:
            if item in highest_items:
                continue
            record_error(
                batch,
                item["line"],
                ValueError(
                    "doublon ignoré : prime TTC inférieure à celle de "
                    f"la ligne {winner_line}"
                ),
            )

    parsed = sorted(retained_by_premium, key=lambda item: item["line"])
    unique_parsed = []
    seen_fingerprints = {}
    seen_keys = {}
    for item in parsed:
        fingerprint = item_fingerprint(item)
        duplicate_line = seen_fingerprints.get(fingerprint)
        if duplicate_line is not None:
            record_error(
                batch,
                item["line"],
                ValueError(f"doublon ignoré : même contrat que la ligne {duplicate_line}"),
            )
            continue
        duplicate_key_item = seen_keys.get(item["key"])
        if (
            duplicate_key_item is not None
            and duplicate_key_item["is_termination"] == item["is_termination"]
        ):
            record_error(
                batch,
                item["line"],
                ValueError(
                    "doublon ignoré : même police et même quittance "
                    f"que la ligne {duplicate_key_item['line']}"
                ),
            )
            continue
        seen_fingerprints[fingerprint] = item["line"]
        seen_keys.setdefault(item["key"], item)
        unique_parsed.append(item)
    parsed = combine_batch_termination_rows(unique_parsed)

    if not parsed:
        return save_batch(batch)

    try:
        with transaction.atomic():
            policies = {
                policy
                for item in parsed
                for policy in (
                    item["policy"],
                    item["legacy_policy"],
                    canonical_policy(
                        item["policy"],
                        item["values"]["category"],
                    ),
                    canonical_policy(
                        item["legacy_policy"],
                        item["values"]["category"],
                    ),
                )
                if policy
            }
            existing_contract_list = list(
                Contract.objects.select_related("client", "termination").filter(
                    policy_number__in=policies
                )
            )
            existing_contracts = {
                (contract.policy_number, contract.receipt): contract
                for contract in existing_contract_list
            }
            existing_contracts_by_id = {
                contract.pk: contract
                for contract in existing_contract_list
            }
            contracts_by_policy = {}
            clients_by_policy = {}
            for contract in existing_contract_list:
                aliases = {
                    contract.policy_number,
                    canonical_policy(contract.policy_number, contract.category),
                }
                for policy in aliases:
                    contracts_by_policy.setdefault(policy, []).append(contract)
                    clients_by_policy.setdefault(policy, contract.client)

            existing_fingerprints = {}
            existing_business_contracts = {}
            for contract in existing_contract_list:
                existing_fingerprints.setdefault(contract_fingerprint(contract), contract)
                identity = stored_contract_business_identity(contract)
                if identity is None:
                    continue
                current = existing_business_contracts.get(identity)
                if (
                    current is None
                    or premium_rank(contract.total_premium)
                    > premium_rank(current.total_premium)
                ):
                    existing_business_contracts[identity] = contract

            filtered_parsed = []
            for item in parsed:
                duplicate_contract = existing_fingerprints.get(item_fingerprint(item))
                exact_contract = existing_contracts.get(item["key"])
                business_contract = existing_business_contracts.get(
                    item_business_identity(item)
                )
                policy_alias = canonical_policy(
                    item["policy"],
                    item["values"]["category"],
                )
                policy_candidates = indexed_policy_candidates(
                    contracts_by_policy,
                    item["policy"],
                    item["legacy_policy"],
                    policy_alias,
                )
                if item["is_termination"]:
                    termination_candidate = select_termination_candidate(
                        item,
                        policy_candidates,
                    )
                    if termination_candidate is not None:
                        item["matched_contract_id"] = termination_candidate.pk
                        if (
                            hasattr(termination_candidate, "termination")
                            and termination_candidate.termination.date
                            < item["termination_date"]
                        ):
                            item["termination_date"] = (
                                termination_candidate.termination.date
                            )
                            item["termination_reason"] = (
                                termination_candidate.termination.reason
                                or item["termination_reason"]
                            )
                            # Le montant doit rester lié à la première
                            # date d'arrêt. Une ristourne plus tardive ne doit
                            # pas devenir la prime de la résiliation initiale.
                            item["termination_premium"] = (
                                termination_candidate.termination.premium
                            )
                    # Une résiliation doit être appliquée même si sa prime est
                    # négative ou inférieure à celle du contrat en cours.
                    filtered_parsed.append(item)
                    continue
                closed_cycle = select_closed_cycle_candidate(
                    item,
                    policy_candidates,
                )
                if closed_cycle is not None:
                    item["matched_contract_id"] = closed_cycle.pk
                    item["preserve_closed_cycle"] = True
                    filtered_parsed.append(item)
                    continue
                if business_contract is not None:
                    if item["import_type"] == ImportBatch.ImportType.UPCOMING:
                        # Le fichier d'échéances ne contient pas de prime : il
                        # complète toujours le contrat de production existant
                        # au lieu d'être rejeté comme un doublon moins cher.
                        item["matched_contract_id"] = business_contract.pk
                        filtered_parsed.append(item)
                        continue
                    incoming_premium = premium_rank(
                        item["values"]["total_premium"]
                    )
                    stored_premium = premium_rank(
                        business_contract.total_premium
                    )
                    if incoming_premium < stored_premium:
                        record_error(
                            batch,
                            item["line"],
                            ValueError(
                                "doublon ignoré : la base contient déjà "
                                "ce contrat avec une prime TTC supérieure"
                            ),
                        )
                        continue
                    if incoming_premium > stored_premium:
                        item["matched_contract_id"] = business_contract.pk
                        filtered_parsed.append(item)
                        continue
                    if exact_contract is not None:
                        item["matched_contract_id"] = business_contract.pk
                        filtered_parsed.append(item)
                        continue
                is_legacy_match = (
                    duplicate_contract is not None
                    and duplicate_contract.policy_number != item["policy"]
                    and canonical_policy(
                        duplicate_contract.policy_number,
                        duplicate_contract.category,
                    ) == canonical_policy(
                        item["policy"],
                        item["values"]["category"],
                    )
                    and duplicate_contract.receipt == item["receipt"]
                )
                if (
                    business_contract is not None
                    and exact_contract is None
                    and not is_legacy_match
                ):
                    duplicate_contract = business_contract
                if duplicate_contract is not None and exact_contract is None and not is_legacy_match:
                    record_error(
                        batch,
                        item["line"],
                        ValueError(
                            "doublon ignoré : ce contrat existe déjà "
                            f"(police {duplicate_contract.policy_number})"
                        ),
                    )
                    continue
                filtered_parsed.append(item)
            parsed = filtered_parsed

            external_ids = {item["external_id"] for item in parsed if item["external_id"]}
            names = {item["name"].lower() for item in parsed}
            clients_by_identity = {
                ("external", client.external_id): client
                for client in Client.objects.filter(external_id__in=external_ids)
            }
            if names:
                for client in Client.objects.annotate(import_name=Lower("name")).filter(import_name__in=names):
                    clients_by_identity.setdefault(("name", client.name.lower()), client)

            new_clients = {}
            changed_clients = {}
            for item in parsed:
                external_identity = (
                    ("external", item["external_id"])
                    if item["external_id"]
                    else None
                )
                name_identity = ("name", item["name"].lower())
                identity = external_identity or name_identity
                client = (
                    (clients_by_identity.get(external_identity) if external_identity else None)
                    or clients_by_policy.get(item["policy"])
                    or clients_by_identity.get(name_identity)
                    or new_clients.get(identity)
                )
                if client is None:
                    client = Client(
                        name=item["name"],
                        phone=item["phone"],
                        external_id=item["external_id"],
                    )
                    new_clients[identity] = client
                else:
                    if item["phone"] and client.phone != item["phone"]:
                        client.phone = item["phone"]
                    if item["external_id"] and not client.external_id:
                        client.external_id = item["external_id"]
                    if client.pk:
                        changed_clients[client.pk] = client
                clients_by_identity[identity] = client
                clients_by_identity.setdefault(name_identity, client)
                if external_identity:
                    clients_by_identity[external_identity] = client
                clients_by_policy.setdefault(item["policy"], client)

            if new_clients:
                Client.objects.bulk_create(list(new_clients.values()), batch_size=500)
            if changed_clients:
                now = timezone.now()
                for client in changed_clients.values():
                    client.updated_at = now
                Client.objects.bulk_update(
                    list(changed_clients.values()),
                    ["phone", "external_id", "updated_at"],
                    batch_size=500,
                )

            final_records = {}
            for item in parsed:
                final_records[item["key"]] = item

            new_contracts = []
            changed_contracts = []
            terminated_contracts = []
            claimed_ids = set()
            now = timezone.now()
            for key, item in final_records.items():
                identity = (
                    ("external", item["external_id"])
                    if item["external_id"]
                    else ("name", item["name"].lower())
                )
                contract = existing_contracts_by_id.get(
                    item.get("matched_contract_id")
                ) or existing_contracts.get(key)
                if contract is None:
                    policy_alias = canonical_policy(
                        item["policy"],
                        item["values"]["category"],
                    )
                    contract = select_contract_candidate(
                        item,
                        indexed_policy_candidates(
                            contracts_by_policy,
                            item["policy"],
                            item["legacy_policy"],
                            policy_alias,
                        ),
                        claimed_ids,
                    )
                if contract is None:
                    contract_values = dict(item["values"])
                    if item["is_termination"]:
                        contract_values["end_date"] = item["termination_date"]
                    contract = Contract(
                        client=clients_by_identity[identity],
                        policy_number=item["policy"],
                        receipt=item["receipt"],
                        from_upcoming_file=(
                            item["import_type"] == ImportBatch.ImportType.UPCOMING
                        ),
                        **contract_values,
                    )
                    new_contracts.append(contract)
                    batch.added_rows += 1
                else:
                    claimed_ids.add(contract.pk)
                    if item.get("preserve_closed_cycle"):
                        merge_closed_cycle_values(
                            contract,
                            item,
                            clients_by_identity[identity],
                        )
                    elif item["is_termination"]:
                        merge_termination_values(
                            contract,
                            item,
                            clients_by_identity[identity],
                        )
                    else:
                        merge_contract_values(
                            contract,
                            item,
                            clients_by_identity[identity],
                        )
                    contract.updated_at = now
                    changed_contracts.append(contract)
                    batch.updated_rows += 1
                if item["is_termination"]:
                    contract.renewal_status = Contract.RenewalStatus.TERMINATED
                    contract.renewed_contract = None
                    terminated_contracts.append((contract, item))
                elif item.get("preserve_closed_cycle"):
                    pass
                elif item["renewed"] and not contract.manually_terminated:
                    contract.renewal_status = Contract.RenewalStatus.RENEWED

            if new_contracts:
                Contract.objects.bulk_create(new_contracts, batch_size=500)

            termination_by_contract = {}
            for contract, item in terminated_contracts:
                current = termination_by_contract.get(contract.pk)
                if current is None or (
                    item["termination_date"],
                    item["line"],
                ) < (
                    current[1]["termination_date"],
                    current[1]["line"],
                ):
                    termination_by_contract[contract.pk] = (contract, item)
            terminated_contracts = list(termination_by_contract.values())
            for contract, item in terminated_contracts:
                contract.end_date = item["termination_date"]
                contract.event = item["termination_reason"]
                contract.renewal_status = Contract.RenewalStatus.TERMINATED
                contract.renewed_contract = None
                contract.updated_at = now
                changed_contracts.append(contract)

            if changed_contracts:
                changed_contracts = list({
                    contract.pk: contract for contract in changed_contracts
                }.values())
                Contract.objects.bulk_update(
                    changed_contracts,
                    [
                        "client",
                        "policy_number",
                        "receipt",
                        *CONTRACT_VALUE_FIELDS,
                        "from_upcoming_file",
                        "renewal_status",
                        "renewed_contract",
                        "updated_at",
                    ],
                    batch_size=500,
                )

            terminated_ids = [
                contract.pk for contract, _item in terminated_contracts
            ]
            existing_terminations = {
                termination.contract_id: termination
                for termination in Termination.objects.filter(contract_id__in=terminated_ids)
            }
            new_terminations = []
            changed_terminations = []
            for contract, item in terminated_contracts:
                termination_date = item["termination_date"]
                termination_reason = item["termination_reason"]
                termination_premium = item["termination_premium"]
                termination = existing_terminations.get(contract.pk)
                if termination is None:
                    new_terminations.append(
                        Termination(
                            contract=contract,
                            date=termination_date,
                            reason=termination_reason,
                            premium=termination_premium,
                            recorded_by=user,
                        )
                    )
                else:
                    changed = False
                    if termination.date != termination_date:
                        termination.date = termination_date
                        changed = True
                    if termination.reason != termination_reason:
                        termination.reason = termination_reason
                        changed = True
                    if (
                        termination_premium is not None
                        and termination.premium != termination_premium
                    ):
                        termination.premium = termination_premium
                        changed = True
                    if changed:
                        changed_terminations.append(termination)
            Termination.objects.bulk_create(new_terminations, batch_size=500)
            if changed_terminations:
                Termination.objects.bulk_update(
                    changed_terminations,
                    ["date", "reason", "premium"],
                    batch_size=500,
                )
            mark_vehicle_renewals()
    except Exception as exc:
        batch.added_rows = 0
        batch.updated_rows = 0
        batch.rejected_rows += len(parsed)
        if len(batch.errors) < 100:
            batch.errors.append({"line": 1, "error": f"Import annulé : {exc}"})

    return save_batch(batch)


def import_contracts(upload, user, expected_type=None):
    rows = read_rows(upload)
    detected_type = detect_import_type(header_map(rows[0])) if rows else ImportBatch.ImportType.GENERAL
    if expected_type and detected_type != expected_type:
        labels = dict(ImportBatch.ImportType.choices)
        detected_label = labels.get(detected_type, "Fichier Excel standard")
        expected_label = labels.get(expected_type, "ce type de fichier")
        raise ValueError(
            f"Ce fichier est détecté comme « {detected_label} ». "
            f"Déposez-le dans la case « {detected_label} » et non dans "
            f"« {expected_label} »."
        )
    return import_contract_rows(rows, upload.name, user)
